"""链接处理测试模块"""
import pytest
import sys
import os
from unittest.mock import Mock, patch, mock_open
import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from dingtalk_download.link_handler import (
    read_links_file,
    extract_live_uuid,
    _get_file_extension,
    _read_csv_links,
    _read_excel_links,
    _extract_links_from_dataframe
)


class TestGetFileExtension:
    """测试 _get_file_extension 函数"""

    def test_get_csv_extension(self):
        """测试获取CSV文件扩展名"""
        result = _get_file_extension("test.csv")
        assert result == ".csv"

    def test_get_xlsx_extension(self):
        """测试获取XLSX文件扩展名"""
        result = _get_file_extension("test.xlsx")
        assert result == ".xlsx"

    def test_get_xls_extension(self):
        """测试获取XLS文件扩展名"""
        result = _get_file_extension("test.xls")
        assert result == ".xls"

    def test_get_uppercase_extension(self):
        """测试获取大写扩展名"""
        result = _get_file_extension("test.CSV")
        assert result == ".csv"

    def test_get_extension_with_path(self):
        """测试获取带路径的文件扩展名"""
        result = _get_file_extension("/path/to/test.csv")
        assert result == ".csv"

    def test_get_extension_no_extension(self):
        """测试没有扩展名的文件"""
        result = _get_file_extension("testfile")
        assert result == ""


class TestReadCsvLinks:
    """测试 _read_csv_links 函数"""

    def test_read_csv_success(self, tmp_path):
        """测试成功读取CSV文件"""
        csv_content = "url\nhttps://n.dingtalk.com/live?liveUuid=test123"
        file_path = tmp_path / "links.csv"
        file_path.write_text(csv_content, encoding='utf-8')

        result = _read_csv_links(str(file_path))
        assert len(result) == 1

    def test_read_csv_gbk_encoding(self, tmp_path):
        """测试使用GBK编码读取CSV文件"""
        csv_content = "url\nhttps://n.dingtalk.com/live?liveUuid=test123"
        file_path = tmp_path / "links.csv"
        file_path.write_text(csv_content, encoding='gbk')

        result = _read_csv_links(str(file_path))
        assert len(result) == 1

    def test_read_csv_utf8_sig_encoding(self, tmp_path):
        """测试使用UTF-8-SIG编码读取CSV文件"""
        csv_content = "\ufeffurl\nhttps://n.dingtalk.com/live?liveUuid=test123"
        file_path = tmp_path / "links.csv"
        file_path.write_text(csv_content, encoding='utf-8-sig')

        result = _read_csv_links(str(file_path))
        assert len(result) == 1

    def test_read_csv_all_encodings_fail(self, tmp_path):
        """测试所有编码都失败"""
        csv_content = "url\nhttps://n.dingtalk.com/live?liveUuid=test123"
        file_path = tmp_path / "links.csv"
        file_path.write_text(csv_content, encoding='utf-16')

        with pytest.raises(RuntimeError) as exc_info:
            _read_csv_links(str(file_path))
        assert "编码无法识别" in str(exc_info.value)


class TestReadExcelLinks:
    """测试 _read_excel_links 函数"""

    def test_read_excel_single_sheet(self, tmp_path):
        """测试读取Excel文件单个工作表"""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'url'
            ws['A2'] = 'https://n.dingtalk.com/live?liveUuid=test123'

            file_path = tmp_path / "links.xlsx"
            wb.save(str(file_path))

            result = _read_excel_links(str(file_path))
            assert len(result) == 1
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_read_excel_multiple_sheets(self, tmp_path):
        """测试读取Excel文件多个工作表"""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws1 = wb.active
            ws1['A1'] = 'url'
            ws1['A2'] = 'https://n.dingtalk.com/live?liveUuid=test123'
            ws1['A3'] = 'https://n.dingtalk.com/live?liveUuid=test456'

            ws2 = wb.create_sheet("Sheet2")
            ws2['A1'] = 'url'
            ws2['A2'] = 'https://n.dingtalk.com/live?liveUuid=abc456'

            file_path = tmp_path / "links.xlsx"
            wb.save(str(file_path))

            result = _read_excel_links(str(file_path))
            assert len(result) >= 1
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_read_excel_error(self, tmp_path):
        """测试读取Excel文件时发生错误"""
        file_path = tmp_path / "invalid.xlsx"
        file_path.write_text("invalid content")

        with pytest.raises(RuntimeError) as exc_info:
            _read_excel_links(str(file_path))
        assert "读取Excel文件时发生错误" in str(exc_info.value)


class TestExtractLinksFromDataframe:
    """测试 _extract_links_from_dataframe 函数"""

    def test_extract_links_single_column(self):
        """测试从单列DataFrame提取链接"""
        df = pd.DataFrame({
            'url': [
                'https://n.dingtalk.com/live?liveUuid=test123',
                'https://example.com/not-dingtalk',
                'https://n.dingtalk.com/live?liveUuid=abc456'
            ]
        })

        result = _extract_links_from_dataframe(df)
        assert len(result) == 2
        assert result[0] == 'https://n.dingtalk.com/live?liveUuid=test123'
        assert result[2] == 'https://n.dingtalk.com/live?liveUuid=abc456'

    def test_extract_links_multiple_columns(self):
        """测试从多列DataFrame提取链接"""
        df = pd.DataFrame({
            'url': ['https://n.dingtalk.com/live?liveUuid=test123', 'https://example.com/not-dingtalk'],
            'name': ['Test Video', 'Test Video 2'],
            'link': ['https://example.com/not-dingtalk', 'https://n.dingtalk.com/live?liveUuid=abc456']
        })

        result = _extract_links_from_dataframe(df)
        assert len(result) == 2

    def test_extract_links_with_nan(self):
        """测试处理包含NaN的DataFrame"""
        df = pd.DataFrame({
            'url': [
                'https://n.dingtalk.com/live?liveUuid=test123',
                None,
                'https://n.dingtalk.com/live?liveUuid=abc456'
            ]
        })

        result = _extract_links_from_dataframe(df)
        assert len(result) == 2

    def test_extract_links_no_dingtalk_links(self):
        """测试没有钉钉链接的DataFrame"""
        df = pd.DataFrame({
            'url': ['https://example.com/video1', 'https://example.com/video2']
        })

        result = _extract_links_from_dataframe(df)
        assert len(result) == 0

    def test_extract_links_empty_dataframe(self):
        """测试空DataFrame"""
        df = pd.DataFrame()

        result = _extract_links_from_dataframe(df)
        assert len(result) == 0


class TestReadLinksFile:
    """测试read_links_file函数"""
    
    def test_read_csv_with_valid_links(self, tmp_path):
        """测试读取包含有效链接的CSV文件"""
        csv_content = "url\nhttps://n.dingtalk.com/live?liveUuid=test123\nhttps://n.dingtalk.com/live?liveUuid=abc456"
        file_path = tmp_path / "links.csv"
        file_path.write_text(csv_content, encoding='utf-8')
        
        result = read_links_file(str(file_path))
        assert len(result) == 2
        assert any("test123" in url for url in result.values())
        assert any("abc456" in url for url in result.values())
    
    def test_read_csv_with_gbk_encoding(self, tmp_path):
        """测试读取GBK编码的CSV文件"""
        csv_content = "url\nhttps://n.dingtalk.com/live?liveUuid=test123"
        file_path = tmp_path / "links.csv"
        file_path.write_text(csv_content, encoding='gbk')
        
        result = read_links_file(str(file_path))
        assert len(result) == 1
    
    def test_read_excel_file(self, tmp_path):
        """测试读取Excel文件"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'url'
            ws['A2'] = 'https://n.dingtalk.com/live?liveUuid=test123'
            
            file_path = tmp_path / "links.xlsx"
            wb.save(str(file_path))
            
            result = read_links_file(str(file_path))
            assert len(result) == 1
        except ImportError:
            pytest.skip("openpyxl not installed")
    
    def test_invalid_file_format_raises_systemexit(self, tmp_path):
        """测试不支持的文件格式抛出 ValueError"""
        file_path = tmp_path / "links.txt"
        file_path.write_text("some content")
        
        with pytest.raises(ValueError) as excinfo:
            read_links_file(str(file_path))
        assert "文件格式不支持" in str(excinfo.value)
    
    def test_no_valid_links_raises_systemexit(self, tmp_path):
        """测试没有有效链接时抛出 ValueError"""
        csv_content = "url\nhttps://example.com/not-dingtalk"
        file_path = tmp_path / "links.csv"
        file_path.write_text(csv_content, encoding='utf-8')
        
        with pytest.raises(ValueError) as excinfo:
            read_links_file(str(file_path))
        assert "未找到有效的钉钉直播链接" in str(excinfo.value)


class TestExtractLiveUuid:
    """测试extract_live_uuid函数"""
    
    def test_extract_from_valid_url(self):
        """测试从有效URL提取liveUuid"""
        url = "https://n.dingtalk.com/live?liveUuid=abc123-def456"
        result = extract_live_uuid(url)
        assert result == "abc123-def456"
    
    def test_returns_none_for_missing_liveUuid(self):
        """测试没有liveUuid时返回None"""
        url = "https://n.dingtalk.com/live"
        result = extract_live_uuid(url)
        assert result is None
